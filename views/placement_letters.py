"""Placement Letters page — client and candidate confirmation letters."""

import streamlit as st
from datetime import datetime
from ui import page_header, step_flow, form_section, convert_docx_to_pdf, build_files_dict


def render():
    page_header("Placement Letters", "Client and candidate placement confirmation letters")

    has_generated = "pl_generated" in st.session_state
    current_step = 1 if has_generated else 0
    step_flow(["Fill Details", "Download & Send"], current_step)

    if not has_generated:
        _render_form()
    else:
        _render_results()


def _render_form():
    from generators.placement_letters import generate_client_letter, generate_candidate_letter

    # --- Spreadsheet Upload ---
    with st.expander("Upload Candidate Details.xlsx (optional — pre-fills the form)"):
        uploaded_xlsx = st.file_uploader(
            "Choose file",
            type=["xlsx"],
            key="pl_upload",
            label_visibility="collapsed",
        )

    if uploaded_xlsx and "pl_xlsx_parsed" not in st.session_state:
        try:
            _parse_spreadsheet(uploaded_xlsx)
            st.rerun()
        except Exception as e:
            st.error(f"Error reading spreadsheet: {e}")

    # --- Form ---
    form_section("Consultant & Candidate")
    col1, col2 = st.columns(2)
    consultant_options = ["Jason Beith", "Tate McClenaghan", "Kelsi Flynn"]
    consultant_idx = st.session_state.get("pl_consultant_idx", 1)

    with col1:
        pl_consultant = st.selectbox("Consultant", consultant_options, index=consultant_idx)
        pl_candidate_name = st.text_input("Candidate name *", key="pl_candidate")
        pl_candidate_address = st.text_area(
            "Candidate address", height=80, key="pl_candidate_addr",
            placeholder="123 Queen St\nAuckland 1010",
        )
    with col2:
        pl_position = st.text_input("Position title *", key="pl_position")
        pl_salary = st.text_input("Salary / package *", key="pl_salary", placeholder="$250,000 + KiwiSaver")
        pl_hiring_manager = st.text_input("Reporting manager", key="pl_manager")

    form_section("Client Details")
    col3, col4 = st.columns(2)
    with col3:
        pl_client_company = st.text_input("Client company *", key="pl_company")
        pl_client_contact = st.text_input("Client contact name *", key="pl_contact")
    with col4:
        pl_client_title = st.text_input("Client contact title", key="pl_contact_title")
        pl_client_address = st.text_area(
            "Client address", height=80, key="pl_client_addr",
            placeholder="456 Shortland St\nAuckland 1010",
        )

    form_section("Placement Details")
    col5, col6 = st.columns(2)
    with col5:
        pl_start_date = st.date_input("Start date", key="pl_date")
        pl_location = st.text_input("Location of work", value="As agreed", key="pl_location")
    with col6:
        pl_guarantee = st.text_input(
            "Guarantee period (client letter only)", value="3 months", key="pl_guarantee",
        )

    form_section("Output")
    lcol1, lcol2 = st.columns(2)
    with lcol1:
        st.markdown("**Which letters?**")
        pl_gen_client = st.checkbox("Client Confirmation", value=True, key="pl_gen_client")
        pl_gen_candidate = st.checkbox("Candidate Confirmation", value=True, key="pl_gen_candidate")
    with lcol2:
        st.markdown("**Format**")
        pl_fmt_docx = st.checkbox(".docx", value=True, key="pl_fmt_docx")
        pl_fmt_pdf = st.checkbox(".pdf", value=False, key="pl_fmt_pdf")

    st.markdown("")
    col_btn, _ = st.columns([1, 2])
    with col_btn:
        generate = st.button("Generate Letters", type="primary", key="pl_generate", use_container_width=True)

    if generate:
        missing = []
        if not pl_candidate_name:
            missing.append("Candidate name")
        if not pl_position:
            missing.append("Position title")
        if not pl_client_company:
            missing.append("Client company")
        if not pl_client_contact:
            missing.append("Client contact name")
        if not pl_salary:
            missing.append("Salary / package")
        if not pl_gen_client and not pl_gen_candidate:
            missing.append("At least one letter type")
        if not pl_fmt_docx and not pl_fmt_pdf:
            missing.append("At least one format")

        if missing:
            st.error(f"Please fill in: {', '.join(missing)}")
        else:
            formatted_date = f"{pl_start_date.day} {pl_start_date.strftime('%B')} {pl_start_date.year}"
            letter_date = f"{datetime.now().day} {datetime.now().strftime('%B')} {datetime.now().year}"

            data = {
                "consultant": pl_consultant,
                "candidate_name": pl_candidate_name,
                "candidate_address": pl_candidate_address,
                "position": pl_position,
                "client_company": pl_client_company,
                "client_contact_name": pl_client_contact,
                "client_contact_title": pl_client_title,
                "client_address": pl_client_address,
                "start_date": formatted_date,
                "salary": pl_salary,
                "hiring_manager": pl_hiring_manager,
                "location_of_work": pl_location,
                "guarantee_period": pl_guarantee,
                "letter_date": letter_date,
            }

            generated = {}
            try:
                if pl_gen_client:
                    generated["client"] = generate_client_letter(data)
                if pl_gen_candidate:
                    generated["candidate"] = generate_candidate_letter(data)
                st.session_state.pl_generated = generated
                st.session_state.pl_data = data
                st.session_state.pl_out_docx = pl_fmt_docx
                st.session_state.pl_out_pdf = pl_fmt_pdf
                st.rerun()
            except Exception as e:
                st.error(f"Error generating letters: {e}")


def _render_results():
    from ms_auth import build_outlook_compose_url

    data = st.session_state.pl_data
    generated = st.session_state.pl_generated
    candidate = data["candidate_name"]
    company = data["client_company"]
    client_first = data["client_contact_name"].split()[0]
    cand_first = candidate.split()[0]
    fmt_docx = st.session_state.get("pl_out_docx", True)
    fmt_pdf = st.session_state.get("pl_out_pdf", False)

    # Back button
    if st.button("< Back to form"):
        del st.session_state["pl_generated"]
        del st.session_state["pl_data"]
        st.rerun()

    # Build files
    name_map = {
        "client": f"{candidate} Placement Confirmation for {company}",
        "candidate": f"Placement Confirmation {candidate} at {company}",
    }
    all_files = build_files_dict(generated, name_map, fmt_docx, fmt_pdf)

    # Split files by letter type
    client_files = {k: v for k, v in all_files.items() if "Confirmation for" in k}
    candidate_files = {k: v for k, v in all_files.items() if "Confirmation for" not in k}

    # --- Download ---
    form_section("Download")
    docx_mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    pdf_mime = "application/pdf"

    for fname, fbytes in all_files.items():
        mime = pdf_mime if fname.endswith(".pdf") else docx_mime
        st.download_button(
            label=f"Download {fname}",
            data=fbytes,
            file_name=fname,
            mime=mime,
            key=f"dl_{fname}",
        )

    if len(all_files) > 1:
        import io, zipfile
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname, fbytes in all_files.items():
                zf.writestr(fname, fbytes)
        st.download_button(
            label="Download All (.zip)",
            data=zip_buffer.getvalue(),
            file_name=f"Placement Letters - {candidate}.zip",
            mime="application/zip",
            type="primary",
            key="dl_all_zip",
        )

    # --- Email ---
    form_section("Email")
    st.caption("Opens Outlook with to/subject/body pre-filled. Attach the downloaded letter to the email.")

    email_cols = st.columns(2)
    if "client" in generated:
        with email_cols[0]:
            client_email = st.text_input(
                "Client email", key="email_client_addr",
                placeholder=f"{client_first.lower()}@company.co.nz",
            )
            if client_email:
                url = build_outlook_compose_url(
                    client_email,
                    f"Placement Confirmation - {candidate}, {data['position']}",
                    f"Dear {client_first},\n\nPlease find attached the placement confirmation for {candidate} as {data['position']} at {company}.\n\nKind regards",
                )
                st.link_button("Email Client Letter", url, type="primary", use_container_width=True)

    if "candidate" in generated:
        with email_cols[1]:
            cand_email = st.text_input("Candidate email", key="email_cand_addr")
            if cand_email:
                url = build_outlook_compose_url(
                    cand_email,
                    f"Congratulations - {data['position']} at {company}",
                    f"Dear {cand_first},\n\nCongratulations on your new role. Please find attached your placement confirmation for {data['position']} at {company}.\n\nKind regards",
                )
                st.link_button("Email Candidate Letter", url, type="primary", use_container_width=True)


def _parse_spreadsheet(uploaded_xlsx):
    """Parse Candidate Details.xlsx and pre-fill session state."""
    from openpyxl import load_workbook
    from datetime import datetime
    import io as _io

    wb = load_workbook(_io.BytesIO(uploaded_xlsx.read()), read_only=True, data_only=True)
    ws = wb.active

    section_headers = {"Candidate", "Role", "Placement", "Referee 1", "Referee 2"}
    field_map = {
        ("Candidate", "Full Name"): "candidate_name",
        ("Candidate", "Address Line 1"): "candidate_address_line_1",
        ("Candidate", "Address Line 2"): "candidate_address_line_2",
        ("Role", "Position"): "position",
        ("Role", "Client Company"): "client_company",
        ("Role", "Client Contact"): "client_contact_name",
        ("Role", "Consultant"): "consultant_raw",
        ("Placement", "Start Date"): "start_date",
        ("Placement", "Salary (Permanent)"): "salary",
        ("Placement", "Pay Rate (Contract)"): "pay_rate",
        ("Placement", "Reporting Manager"): "reporting_manager",
        ("Placement", "Client Address Line 1"): "client_address_line_1",
        ("Placement", "Client Address Line 2"): "client_address_line_2",
    }

    parsed = {}
    current_section = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        label = str(row[0].value or "").strip()
        value = row[1].value
        if isinstance(value, datetime):
            value = f"{value.day} {value.strftime('%B')} {value.year}"
        else:
            value = str(value).strip() if value is not None else ""
        if label in section_headers and not value:
            current_section = label
            continue
        if not label or not current_section:
            continue
        key = (current_section, label)
        if key in field_map:
            parsed[field_map[key]] = value
    wb.close()

    # Resolve consultant
    consultant_map = {
        "jason": "Jason Beith", "jason beith": "Jason Beith",
        "kelsi": "Kelsi Flynn", "kelsi flynn": "Kelsi Flynn",
        "tate": "Tate McClenaghan", "tate mcclenaghan": "Tate McClenaghan",
    }
    raw = parsed.get("consultant_raw", "")
    parsed["consultant"] = consultant_map.get(raw.lower(), "Tate McClenaghan")

    # Combine address lines
    parsed["candidate_address"] = "\n".join(
        l for l in [parsed.get("candidate_address_line_1", ""),
                    parsed.get("candidate_address_line_2", "")] if l
    )
    parsed["client_address"] = "\n".join(
        l for l in [parsed.get("client_address_line_1", ""),
                    parsed.get("client_address_line_2", "")] if l
    )

    if not parsed.get("salary"):
        parsed["salary"] = parsed.get("pay_rate", "")

    # Set widget keys
    st.session_state.pl_candidate = parsed.get("candidate_name", "")
    st.session_state.pl_candidate_addr = parsed.get("candidate_address", "")
    st.session_state.pl_position = parsed.get("position", "")
    st.session_state.pl_salary = parsed.get("salary", "")
    st.session_state.pl_company = parsed.get("client_company", "")
    st.session_state.pl_contact = parsed.get("client_contact_name", "")
    st.session_state.pl_client_addr = parsed.get("client_address", "")
    st.session_state.pl_manager = parsed.get("reporting_manager", "")

    consultant_options = ["Jason Beith", "Tate McClenaghan", "Kelsi Flynn"]
    st.session_state.pl_consultant_idx = (
        consultant_options.index(parsed["consultant"])
        if parsed.get("consultant") in consultant_options
        else 1
    )
    st.session_state.pl_xlsx_parsed = True
