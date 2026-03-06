"""Document Hub - branded document generator for the team."""

import io
import streamlit as st
from datetime import date, datetime


def convert_docx_to_pdf(docx_bytes: bytes, filename: str = "document.docx") -> bytes | None:
    """Convert .docx bytes to .pdf bytes via Microsoft Graph API."""
    from ms_auth import convert_docx_to_pdf_graph
    return convert_docx_to_pdf_graph(docx_bytes, filename)

st.set_page_config(
    page_title="Document Hub",
    page_icon="I",
    layout="wide",
)


# --- Auth Gate ---
from ms_auth import ms_login

if not ms_login():
    st.stop()


# --- Brand Detection ---
from brands import get_brand, get_brand_css

user_email = st.session_state.get("ms_email", "")
brand = get_brand(user_email)
st.markdown(get_brand_css(brand), unsafe_allow_html=True)

# Clean up expired drafts (once per session, silent if Supabase not configured)
if "_drafts_cleaned" not in st.session_state:
    try:
        from drafts import cleanup_expired
        cleanup_expired()
    except Exception:
        pass
    st.session_state["_drafts_cleaned"] = True


# --- Sidebar ---
st.sidebar.image(brand["logo_url"], width="stretch")
st.sidebar.markdown(f"**{st.session_state.get('ms_user', 'User')}**")
st.sidebar.divider()

DOCUMENT_TYPES = {
    "Reference Check": "reference_check",
    "Placement Letters": "placement_letters",
    "Terms & Conditions": "terms_conditions",
    "Contractor Agreement": "contractor_agreement",
    "Assignment Confirmation (coming soon)": None,
    "CV Profile (coming soon)": None,
}

selected = st.sidebar.radio("Document type", list(DOCUMENT_TYPES.keys()), label_visibility="collapsed")

st.sidebar.divider()
if st.sidebar.button("Sign out", width="stretch"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

# Temporary debug — remove once Origin branding is confirmed working
if st.session_state.get("_debug_brand_email"):
    with st.sidebar.expander("Brand debug"):
        st.caption(f"**Email:** {st.session_state.get('_debug_brand_email', '?')}")
        st.caption(f"**Brand:** {brand['short_name']}")
        claims = st.session_state.get("_debug_brand_claims", {})
        for k, v in claims.items():
            st.caption(f"{k}: {v}")

# --- Header ---
st.title("Document Hub")

# --- Reference Check Page ---
if DOCUMENT_TYPES[selected] == "reference_check":
    from generators.reference_check import generate_docx, QUESTIONS
    from ai import process_reference_transcript

    st.header("Reference Check")

    # Form fields
    col1, col2 = st.columns(2)
    with col1:
        candidate_name = st.text_input("Candidate name *")
        position = st.text_input("Role applied for *")
        completed_by = st.selectbox(
            "Completed by",
            ["Tate McClenaghan", "Jason Elston", "Kelsi Halliday", "Aimee"],
        )
    with col2:
        referee_name = st.text_input("Referee name *")
        referee_title = st.text_input("Referee current position")
        referee_previous = st.text_input("Referee previous position (optional)")

    transcript = st.text_area(
        "Paste Granola transcript",
        height=300,
        placeholder="Paste the full reference call transcript here...",
    )

    # --- Generate ---
    if st.button("Generate Reference", type="primary"):
        if not candidate_name or not position or not referee_name:
            st.error("Please fill in all required fields (marked with *).")
        elif not transcript.strip():
            st.error("Please paste the transcript.")
        else:
            with st.spinner("Processing transcript with AI..."):
                try:
                    answers = process_reference_transcript(
                        candidate_name=candidate_name,
                        position=position,
                        referee_name=referee_name,
                        referee_title=referee_title,
                        referee_previous=referee_previous,
                        transcript=transcript,
                    )
                    st.session_state.ref_answers = answers
                    st.session_state.ref_metadata = {
                        "candidate_name": candidate_name,
                        "position": position,
                        "referee_name": referee_name,
                        "referee_title": referee_title,
                        "referee_previous": referee_previous,
                        "completed_by": completed_by,
                        "reference_date": date.today().strftime("%d/%m/%Y"),
                    }
                    st.rerun()
                except Exception as e:
                    st.error(f"Error processing transcript: {e}")

    # --- Review & Edit ---
    if "ref_answers" in st.session_state:
        st.divider()
        st.subheader("Review & Edit Answers")
        st.caption("Edit any answer below before downloading the document.")

        answers = st.session_state.ref_answers
        edited_answers = {}

        for i, question in enumerate(QUESTIONS):
            key = str(i)
            current = answers.get(key, "")
            is_gap = current.startswith("[GAP]")

            label = f"Q{i+1}: {question}"
            if is_gap:
                label += "  ⚠️ NEEDS REVIEW"

            edited = st.text_area(
                label,
                value=current.replace("[GAP] ", ""),
                height=120 if len(current) > 200 else 80,
                key=f"answer_{i}",
            )
            edited_answers[key] = edited

        # --- Download ---
        st.divider()
        metadata = st.session_state.ref_metadata
        data = {**metadata, "answers": edited_answers}

        try:
            docx_bytes = generate_docx(data)
            filename = f"Reference Check for {metadata['candidate_name']} from {metadata['referee_name']}.docx"
            st.download_button(
                label="Download .docx",
                data=docx_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                type="primary",
            )
        except Exception as e:
            st.error(f"Error generating document: {e}")

# --- Placement Letters Page ---
elif DOCUMENT_TYPES.get(selected) == "placement_letters":
    from generators.placement_letters import generate_client_letter, generate_candidate_letter

    st.header("Placement Letters")

    # --- Spreadsheet Upload ---
    uploaded_xlsx = st.file_uploader(
        "Upload Candidate Details.xlsx (optional — pre-fills the form)",
        type=["xlsx"],
        key="pl_upload",
    )

    if uploaded_xlsx and "pl_xlsx_parsed" not in st.session_state:
        try:
            from openpyxl import load_workbook
            import io as _io

            wb = load_workbook(_io.BytesIO(uploaded_xlsx.read()), read_only=True, data_only=True)
            ws = wb.active

            section_headers = {"Candidate", "Role", "Placement", "Referee 1", "Referee 2"}
            field_map = {
                ("Candidate", "Full Name"): "candidate_name",
                ("Candidate", "Address Line 1"): "candidate_address_line_1",
                ("Candidate", "Address Line 2"): "candidate_address_line_2",
                ("Role", "Position"): "position",
                ("Role", "Client Company"): "client_company",
                ("Role", "Client Contact"): "client_contact_name",
                ("Role", "Consultant"): "consultant_raw",
                ("Placement", "Start Date"): "start_date",
                ("Placement", "Salary (Permanent)"): "salary",
                ("Placement", "Pay Rate (Contract)"): "pay_rate",
                ("Placement", "Reporting Manager"): "reporting_manager",
                ("Placement", "Client Address Line 1"): "client_address_line_1",
                ("Placement", "Client Address Line 2"): "client_address_line_2",
            }

            parsed = {}
            current_section = None
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
                label = str(row[0].value or "").strip()
                value = row[1].value
                if isinstance(value, datetime):
                    value = f"{value.day} {value.strftime('%B')} {value.year}"
                else:
                    value = str(value).strip() if value is not None else ""
                if label in section_headers and not value:
                    current_section = label
                    continue
                if not label or not current_section:
                    continue
                key = (current_section, label)
                if key in field_map:
                    parsed[field_map[key]] = value
            wb.close()

            # Resolve consultant name
            consultant_map = {
                "jason": "Jason Beith", "jason beith": "Jason Beith",
                "kelsi": "Kelsi Flynn", "kelsi flynn": "Kelsi Flynn",
                "tate": "Tate McClenaghan", "tate mcclenaghan": "Tate McClenaghan",
            }
            raw_consultant = parsed.get("consultant_raw", "")
            parsed["consultant"] = consultant_map.get(raw_consultant.lower(), "Tate McClenaghan")

            # Combine address lines
            parsed["candidate_address"] = "\n".join(
                l for l in [parsed.get("candidate_address_line_1", ""), parsed.get("candidate_address_line_2", "")] if l
            )
            parsed["client_address"] = "\n".join(
                l for l in [parsed.get("client_address_line_1", ""), parsed.get("client_address_line_2", "")] if l
            )

            # Use salary or pay rate
            if not parsed.get("salary"):
                parsed["salary"] = parsed.get("pay_rate", "")

            # Set widget keys directly in session state so form pre-fills
            st.session_state.pl_candidate = parsed.get("candidate_name", "")
            st.session_state.pl_candidate_addr = parsed.get("candidate_address", "")
            st.session_state.pl_position = parsed.get("position", "")
            st.session_state.pl_salary = parsed.get("salary", "")
            st.session_state.pl_company = parsed.get("client_company", "")
            st.session_state.pl_contact = parsed.get("client_contact_name", "")
            st.session_state.pl_client_addr = parsed.get("client_address", "")
            st.session_state.pl_manager = parsed.get("reporting_manager", "")

            # Consultant selectbox uses index
            consultant_options = ["Jason Beith", "Tate McClenaghan", "Kelsi Flynn"]
            st.session_state.pl_consultant_idx = (
                consultant_options.index(parsed["consultant"])
                if parsed.get("consultant") in consultant_options
                else 1
            )

            st.session_state.pl_xlsx_parsed = True
            st.rerun()
        except Exception as e:
            st.error(f"Error reading spreadsheet: {e}")

    # Form fields
    col1, col2 = st.columns(2)
    consultant_options = ["Jason Beith", "Tate McClenaghan", "Kelsi Flynn"]
    consultant_idx = st.session_state.get("pl_consultant_idx", 1)

    with col1:
        pl_consultant = st.selectbox(
            "Consultant",
            consultant_options,
            index=consultant_idx,
        )
        pl_candidate_name = st.text_input("Candidate name *", key="pl_candidate")
        pl_candidate_address = st.text_area(
            "Candidate address",
            height=80,
            key="pl_candidate_addr",
            placeholder="123 Queen St\nAuckland 1010",
        )
        pl_position = st.text_input("Position title *", key="pl_position")
        pl_salary = st.text_input("Salary / package *", key="pl_salary", placeholder="$250,000 + KiwiSaver")
    with col2:
        pl_client_company = st.text_input("Client company *", key="pl_company")
        pl_client_contact = st.text_input("Client contact name *", key="pl_contact")
        pl_client_title = st.text_input("Client contact title", key="pl_contact_title")
        pl_client_address = st.text_area(
            "Client address",
            height=80,
            key="pl_client_addr",
            placeholder="456 Shortland St\nAuckland 1010",
        )
        pl_hiring_manager = st.text_input("Reporting manager", key="pl_manager")

    col3, col4 = st.columns(2)
    with col3:
        pl_start_date = st.date_input("Start date", key="pl_date")
        pl_location = st.text_input("Location of work", value="As agreed", key="pl_location")
    with col4:
        pl_guarantee = st.text_input("Guarantee period (client letter only)", value="3 months", key="pl_guarantee")

    st.divider()

    # Letter selection
    st.subheader("Output")
    lcol1, lcol2 = st.columns(2)
    with lcol1:
        st.caption("Which letters?")
        pl_gen_client = st.checkbox("Client Confirmation", value=True, key="pl_gen_client")
        pl_gen_candidate = st.checkbox("Candidate Confirmation", value=True, key="pl_gen_candidate")
    with lcol2:
        st.caption("Format")
        pl_fmt_docx = st.checkbox(".docx", value=True, key="pl_fmt_docx")
        pl_fmt_pdf = st.checkbox(".pdf", value=False, key="pl_fmt_pdf")

    # Generate
    if st.button("Generate Letters", type="primary", key="pl_generate"):
        # Validation
        missing = []
        if not pl_candidate_name:
            missing.append("Candidate name")
        if not pl_position:
            missing.append("Position title")
        if not pl_client_company:
            missing.append("Client company")
        if not pl_client_contact:
            missing.append("Client contact name")
        if not pl_salary:
            missing.append("Salary / package")
        if not pl_gen_client and not pl_gen_candidate:
            missing.append("At least one letter type")
        if not pl_fmt_docx and not pl_fmt_pdf:
            missing.append("At least one format")

        if missing:
            st.error(f"Please fill in: {', '.join(missing)}")
        else:
            # Format start date
            formatted_date = f"{pl_start_date.day} {pl_start_date.strftime('%B')} {pl_start_date.year}"
            letter_date = f"{datetime.now().day} {datetime.now().strftime('%B')} {datetime.now().year}"

            data = {
                "consultant": pl_consultant,
                "candidate_name": pl_candidate_name,
                "candidate_address": pl_candidate_address,
                "position": pl_position,
                "client_company": pl_client_company,
                "client_contact_name": pl_client_contact,
                "client_contact_title": pl_client_title,
                "client_address": pl_client_address,
                "start_date": formatted_date,
                "salary": pl_salary,
                "hiring_manager": pl_hiring_manager,
                "location_of_work": pl_location,
                "guarantee_period": pl_guarantee,
                "letter_date": letter_date,
            }

            generated = {}
            try:
                if pl_gen_client:
                    generated["client"] = generate_client_letter(data)
                if pl_gen_candidate:
                    generated["candidate"] = generate_candidate_letter(data)
                st.session_state.pl_generated = generated
                st.session_state.pl_data = data
                st.rerun()
            except Exception as e:
                st.error(f"Error generating letters: {e}")

    # --- Review, Save & Send ---
    if "pl_generated" in st.session_state:
        from ms_auth import build_outlook_compose_url

        st.divider()
        data = st.session_state.pl_data
        candidate = data["candidate_name"]
        company = data["client_company"]
        generated = st.session_state.pl_generated
        client_first = data["client_contact_name"].split()[0]
        cand_first = candidate.split()[0]

        # Build all files
        all_files = {}
        for letter_type, docx_bytes in generated.items():
            if letter_type == "client":
                base_name = f"{candidate} Placement Confirmation for {company}"
            else:
                base_name = f"Placement Confirmation {candidate} at {company}"
            if pl_fmt_docx:
                all_files[f"{base_name}.docx"] = docx_bytes
            if pl_fmt_pdf:
                pdf_bytes = convert_docx_to_pdf(docx_bytes)
                if pdf_bytes:
                    all_files[f"{base_name}.pdf"] = pdf_bytes
                else:
                    st.warning(f"PDF conversion failed for {letter_type} letter.")

        # --- Step 1: Email Preview ---
        st.subheader("1. Email Preview")
        st.caption("Review and edit before sending. Leave email blank to skip.")

        email_cols = st.columns(2)

        with email_cols[0]:
            if "client" in generated:
                st.markdown("**Client Letter**")
                client_email = st.text_input(
                    "To",
                    key="email_client_addr",
                    placeholder=f"{client_first.lower()}@company.co.nz",
                )
                client_subject = st.text_input(
                    "Subject",
                    value=f"Placement Confirmation - {candidate}, {data['position']}",
                    key="email_client_subject",
                )
                client_body = st.text_area(
                    "Email body",
                    value=(
                        f"Dear {client_first},\n\n"
                        f"Please find attached the placement confirmation for "
                        f"{candidate} as {data['position']} at {company}.\n\n"
                        f"Kind regards"
                    ),
                    height=150,
                    key="email_client_body",
                )
            else:
                client_email = ""

        with email_cols[1]:
            if "candidate" in generated:
                st.markdown("**Candidate Letter**")
                cand_email = st.text_input(
                    "To",
                    key="email_cand_addr",
                )
                cand_subject = st.text_input(
                    "Subject",
                    value=f"Congratulations - {data['position']} at {company}",
                    key="email_cand_subject",
                )
                cand_body = st.text_area(
                    "Email body",
                    value=(
                        f"Dear {cand_first},\n\n"
                        f"Congratulations on your new role. Please find attached "
                        f"your placement confirmation for {data['position']} at {company}.\n\n"
                        f"Kind regards"
                    ),
                    height=150,
                    key="email_cand_body",
                )
            else:
                cand_email = ""

        st.divider()

        # Download files
        st.subheader("2. Download")
        docx_mime = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        pdf_mime = "application/pdf"

        for fname, fbytes in all_files.items():
            mime = pdf_mime if fname.endswith(".pdf") else docx_mime
            st.download_button(
                label=f"Download {fname}",
                data=fbytes,
                file_name=fname,
                mime=mime,
                key=f"dl_{fname}",
            )

        if len(all_files) > 1:
            import zipfile
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, fbytes in all_files.items():
                    zf.writestr(fname, fbytes)
            st.download_button(
                label=f"Download All (.zip)",
                data=zip_buffer.getvalue(),
                file_name=f"Placement Letters - {candidate}.zip",
                mime="application/zip",
                type="primary",
                key="dl_all_zip",
            )

        # Outlook compose links
        st.divider()
        st.subheader("3. Send via Outlook")
        st.caption("Opens Outlook with your signature. Attach the downloaded files.")

        email_btns = st.columns(2)
        if client_email:
            with email_btns[0]:
                url = build_outlook_compose_url(client_email, client_subject, client_body)
                st.link_button("Open Client Email in Outlook", url)
        if cand_email:
            with email_btns[1]:
                url = build_outlook_compose_url(cand_email, cand_subject, cand_body)
                st.link_button("Open Candidate Email in Outlook", url)

# --- Terms & Conditions Page ---
elif DOCUMENT_TYPES.get(selected) == "terms_conditions":
    from generators.terms_conditions import generate_docx as generate_tc
    from drafts import save_draft, load_draft, delete_draft

    st.header("Terms & Conditions")

    _DOC_TYPE = "terms_conditions"

    # --- Draft Resume ---
    if f"_{_DOC_TYPE}_checked" not in st.session_state:
        _draft = load_draft(user_email, _DOC_TYPE)
        st.session_state[f"_{_DOC_TYPE}_checked"] = True
        if _draft:
            st.session_state[f"_{_DOC_TYPE}_pending"] = _draft

    if f"_{_DOC_TYPE}_pending" in st.session_state:
        _d = st.session_state[f"_{_DOC_TYPE}_pending"]
        _updated = _d.get("updated_at", "")[:10]
        st.info(f"You have a saved draft from {_updated}.")
        _rc1, _rc2 = st.columns(2)
        with _rc1:
            if st.button("Resume draft", type="primary", key="tc_resume"):
                _fd = _d.get("form_data", {})
                st.session_state["tc_client_name"] = _fd.get("client_name", "")
                if _fd.get("date"):
                    try:
                        st.session_state["tc_date"] = date.fromisoformat(_fd["date"])
                    except ValueError:
                        pass
                st.session_state["tc_guarantee"] = _fd.get("guarantee", 3)
                st.session_state["tc_perm_enabled"] = _fd.get("perm_enabled", True)
                st.session_state["tc_contract_enabled"] = _fd.get("contract_enabled", False)
                st.session_state["tc_exec_enabled"] = _fd.get("exec_enabled", False)
                st.session_state["tc_perm_fee_pct"] = _fd.get("perm_fee_pct", 18)
                st.session_state["tc_perm_basis"] = _fd.get("perm_basis", "Total Salary Package")
                st.session_state["tc_perm_structure"] = _fd.get("perm_structure", "Retained (thirds)")
                st.session_state["tc_perm_fixed_fee"] = _fd.get("perm_fixed_fee", "")
                st.session_state["tc_contract_margin"] = _fd.get("contract_margin", 25)
                st.session_state["tc_exec_fee_pct"] = _fd.get("exec_fee_pct", 25)
                st.session_state["tc_exec_basis"] = _fd.get("exec_basis", "Total Salary Package")
                st.session_state["tc_exec_structure"] = _fd.get("exec_structure", "Retained (thirds)")
                st.session_state["tc_exec_fixed_fee"] = _fd.get("exec_fixed_fee", "")
                st.session_state["tc_sig_infinitas"] = _fd.get("sig_infinitas", False)
                st.session_state["tc_sig_client"] = _fd.get("sig_client", False)
                del st.session_state[f"_{_DOC_TYPE}_pending"]
                st.rerun()
        with _rc2:
            if st.button("Start fresh", key="tc_fresh"):
                del st.session_state[f"_{_DOC_TYPE}_pending"]
                st.rerun()
        st.stop()

    # --- Form ---
    col1, col2 = st.columns(2)
    with col1:
        tc_client = st.text_input("Client company name *", key="tc_client_name")
        tc_date_val = st.date_input("Date", value=date.today(), key="tc_date")
    with col2:
        tc_guarantee = st.selectbox(
            "Guarantee period", [3, 6, 12],
            format_func=lambda x: f"{x} months", key="tc_guarantee",
        )

    st.divider()
    st.subheader("Service Types")

    svc1, svc2, svc3 = st.columns(3)
    with svc1:
        tc_perm = st.checkbox("Permanent / Fixed Term", value=True, key="tc_perm_enabled")
    with svc2:
        tc_contract = st.checkbox("Contractor / Temporary Worker", key="tc_contract_enabled")
    with svc3:
        tc_exec = st.checkbox("Retained / Executive Search", key="tc_exec_enabled")

    if tc_perm:
        st.markdown("**Permanent / Fixed Term Fees**")
        p1, p2, p3 = st.columns(3)
        with p1:
            st.number_input("Fee %", value=18, min_value=1, max_value=100, key="tc_perm_fee_pct")
        with p2:
            st.selectbox("Calculated on", ["Total Salary Package", "Base Salary"], key="tc_perm_basis")
        with p3:
            st.selectbox("Fee structure", ["Retained (thirds)", "Contingent", "Fixed Fee"], key="tc_perm_structure")
        if st.session_state.get("tc_perm_structure") == "Fixed Fee":
            st.text_input("Fixed fee amount", key="tc_perm_fixed_fee")

    if tc_contract:
        st.markdown("**Contractor / Temporary Worker Fees**")
        st.number_input("Margin %", value=25, min_value=1, max_value=100, key="tc_contract_margin")

    if tc_exec:
        st.markdown("**Executive Search Fees**")
        e1, e2, e3 = st.columns(3)
        with e1:
            st.number_input("Fee %", value=25, min_value=1, max_value=100, key="tc_exec_fee_pct")
        with e2:
            st.selectbox("Calculated on ", ["Total Salary Package", "Base Salary"], key="tc_exec_basis")
        with e3:
            st.selectbox("Fee structure ", ["Retained (thirds)", "Contingent", "Fixed Fee"], key="tc_exec_structure")
        if st.session_state.get("tc_exec_structure") == "Fixed Fee":
            st.text_input("Fixed fee amount ", key="tc_exec_fixed_fee")

    st.divider()
    st.subheader("Signature Blocks")
    sig1, sig2 = st.columns(2)
    with sig1:
        st.checkbox("Include Infinitas signature", key="tc_sig_infinitas")
    with sig2:
        st.checkbox("Include Client signature", key="tc_sig_client")

    st.divider()
    st.subheader("Output")
    fmt1, fmt2 = st.columns(2)
    with fmt1:
        tc_fmt_docx = st.checkbox(".docx", value=True, key="tc_fmt_docx")
        tc_fmt_pdf = st.checkbox(".pdf", key="tc_fmt_pdf")

    # Auto-save draft (runs on every rerun = every widget change)
    if tc_client and "tc_generated" not in st.session_state:
        _draft_data = {
            "client_name": tc_client,
            "date": str(st.session_state.get("tc_date", date.today())),
            "guarantee": st.session_state.get("tc_guarantee", 3),
            "perm_enabled": st.session_state.get("tc_perm_enabled", True),
            "contract_enabled": st.session_state.get("tc_contract_enabled", False),
            "exec_enabled": st.session_state.get("tc_exec_enabled", False),
            "perm_fee_pct": st.session_state.get("tc_perm_fee_pct", 18),
            "perm_basis": st.session_state.get("tc_perm_basis", "Total Salary Package"),
            "perm_structure": st.session_state.get("tc_perm_structure", "Retained (thirds)"),
            "perm_fixed_fee": st.session_state.get("tc_perm_fixed_fee", ""),
            "contract_margin": st.session_state.get("tc_contract_margin", 25),
            "exec_fee_pct": st.session_state.get("tc_exec_fee_pct", 25),
            "exec_basis": st.session_state.get("tc_exec_basis", "Total Salary Package"),
            "exec_structure": st.session_state.get("tc_exec_structure", "Retained (thirds)"),
            "exec_fixed_fee": st.session_state.get("tc_exec_fixed_fee", ""),
            "sig_infinitas": st.session_state.get("tc_sig_infinitas", False),
            "sig_client": st.session_state.get("tc_sig_client", False),
        }
        save_draft(user_email, _DOC_TYPE, _draft_data)

    # Generate
    if st.button("Generate T&Cs", type="primary", key="tc_generate"):
        if not tc_client:
            st.error("Please enter the client company name.")
        elif not tc_perm and not tc_contract and not tc_exec:
            st.error("Please enable at least one service type.")
        elif not tc_fmt_docx and not tc_fmt_pdf:
            st.error("Please select at least one output format.")
        else:
            _structure_map = {
                "Retained (thirds)": "retained",
                "Contingent": "contingent",
                "Fixed Fee": "fixed_fee",
            }
            _gen_data = {
                "client_name": tc_client,
                "date": str(st.session_state.get("tc_date", date.today())),
                "perm_enabled": st.session_state.get("tc_perm_enabled", True),
                "contract_enabled": st.session_state.get("tc_contract_enabled", False),
                "exec_enabled": st.session_state.get("tc_exec_enabled", False),
                "perm_fee_pct": st.session_state.get("tc_perm_fee_pct", 18),
                "perm_basis": st.session_state.get("tc_perm_basis", "Total Salary Package").lower(),
                "perm_structure": _structure_map.get(
                    st.session_state.get("tc_perm_structure", "Retained (thirds)"), "retained"
                ),
                "perm_fixed_fee": st.session_state.get("tc_perm_fixed_fee", ""),
                "contract_margin_pct": st.session_state.get("tc_contract_margin", 25),
                "exec_fee_pct": st.session_state.get("tc_exec_fee_pct", 25),
                "exec_basis": st.session_state.get("tc_exec_basis", "Total Salary Package").lower(),
                "exec_structure": _structure_map.get(
                    st.session_state.get("tc_exec_structure", "Retained (thirds)"), "retained"
                ),
                "exec_fixed_fee": st.session_state.get("tc_exec_fixed_fee", ""),
                "guarantee_months": st.session_state.get("tc_guarantee", 3),
                "sig_infinitas": st.session_state.get("tc_sig_infinitas", False),
                "sig_client": st.session_state.get("tc_sig_client", False),
                "adobe_sign": False,
            }
            try:
                docx_bytes = generate_tc(_gen_data)
                st.session_state["tc_generated"] = docx_bytes
                st.session_state["tc_gen_data"] = _gen_data
                delete_draft(user_email, _DOC_TYPE)
                st.rerun()
            except Exception as e:
                st.error(f"Error generating document: {e}")

    # --- Download ---
    if "tc_generated" in st.session_state:
        st.divider()
        st.subheader("Download")
        _gen_data = st.session_state["tc_gen_data"]
        _docx = st.session_state["tc_generated"]
        _fname = f"Infinitas Talent - Terms and Conditions - {_gen_data['client_name']}"

        if tc_fmt_docx:
            st.download_button(
                "Download .docx", _docx, f"{_fname}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="tc_dl_docx",
            )
        if tc_fmt_pdf:
            _pdf = convert_docx_to_pdf(_docx)
            if _pdf:
                st.download_button(
                    "Download .pdf", _pdf, f"{_fname}.pdf",
                    mime="application/pdf",
                    key="tc_dl_pdf",
                )
            else:
                st.warning("PDF conversion failed.")

# --- Contractor Agreement Page ---
elif DOCUMENT_TYPES.get(selected) == "contractor_agreement":
    from generators.contractor_agreement import generate_docx as generate_ca
    from drafts import save_draft, load_draft, delete_draft

    st.header("Contractor Agreement")

    _DOC_TYPE = "contractor_agreement"

    # --- Draft Resume ---
    if f"_{_DOC_TYPE}_checked" not in st.session_state:
        _draft = load_draft(user_email, _DOC_TYPE)
        st.session_state[f"_{_DOC_TYPE}_checked"] = True
        if _draft:
            st.session_state[f"_{_DOC_TYPE}_pending"] = _draft

    if f"_{_DOC_TYPE}_pending" in st.session_state:
        _d = st.session_state[f"_{_DOC_TYPE}_pending"]
        _updated = _d.get("updated_at", "")[:10]
        st.info(f"You have a saved draft from {_updated}.")
        _rc1, _rc2 = st.columns(2)
        with _rc1:
            if st.button("Resume draft", type="primary", key="ca_resume"):
                _fd = _d.get("form_data", {})
                _ctype = _fd.get("type", "sole_trader")
                st.session_state["ca_type"] = "Limited Company" if _ctype == "ltd_company" else "Sole Trader"
                st.session_state["ca_nominated_client"] = _fd.get("nominated_client", "")
                st.session_state["ca_role"] = _fd.get("role", "")
                for _dk in ("commencement_date", "end_date"):
                    if _fd.get(_dk):
                        try:
                            st.session_state[f"ca_{_dk}"] = date.fromisoformat(_fd[_dk])
                        except ValueError:
                            pass
                st.session_state["ca_hours_of_work"] = _fd.get("hours_of_work", "")
                st.session_state["ca_contract_rate"] = _fd.get("contract_rate", "")
                st.session_state["ca_notice_period"] = _fd.get("notice_period", "")
                st.session_state["ca_travel_expenses"] = _fd.get("travel_expenses", "Upon authorisation by the Nominated Client")
                if _ctype == "ltd_company":
                    st.session_state["ca_provider_company"] = _fd.get("provider_company", "")
                    st.session_state["ca_trading_as"] = _fd.get("trading_as", "")
                    st.session_state["ca_registered_address"] = _fd.get("registered_address", "")
                    st.session_state["ca_company_nzbn"] = _fd.get("company_nzbn", "")
                    st.session_state["ca_individual_contractor"] = _fd.get("individual_contractor", "")
                    st.session_state["ca_ird_number"] = _fd.get("ird_number", "")
                    st.session_state["ca_gst_registered"] = _fd.get("gst_registered", "No")
                    st.session_state["ca_gst_number"] = _fd.get("gst_number", "")
                    st.session_state["ca_bank_account"] = _fd.get("bank_account", "")
                del st.session_state[f"_{_DOC_TYPE}_pending"]
                st.rerun()
        with _rc2:
            if st.button("Start fresh", key="ca_fresh"):
                del st.session_state[f"_{_DOC_TYPE}_pending"]
                st.rerun()
        st.stop()

    # --- Form ---
    ca_type = st.radio(
        "Contractor type",
        ["Sole Trader", "Limited Company"],
        key="ca_type",
        horizontal=True,
    )
    is_ltd = ca_type == "Limited Company"

    st.divider()
    st.subheader("Schedule 1 — Assignment Details")

    # Ltd Company extra fields
    if is_ltd:
        st.markdown("**Company Details**")
        l1, l2 = st.columns(2)
        with l1:
            st.text_input("Provider company name *", key="ca_provider_company")
            st.text_input("Trading as (if applicable)", key="ca_trading_as")
            st.text_input("Company No. / NZBN", key="ca_company_nzbn")
        with l2:
            st.text_area("Registered address", height=80, key="ca_registered_address")
            st.text_input("Name of Individual Contractor *", key="ca_individual_contractor")
            st.text_input("IRD Number", key="ca_ird_number")

        g1, g2 = st.columns(2)
        with g1:
            ca_gst = st.radio("GST Registered", ["No", "Yes"], key="ca_gst_registered", horizontal=True)
        with g2:
            if ca_gst == "Yes":
                st.text_input("GST Number", key="ca_gst_number")
        st.text_input("Nominated Bank Account Number", key="ca_bank_account")
        st.divider()

    # Common fields
    st.markdown("**Assignment Details**")
    c1, c2 = st.columns(2)
    with c1:
        ca_client = st.text_input("Nominated Client *", key="ca_nominated_client")
        st.text_input("Role *", key="ca_role")
        st.date_input("Commencement Date", key="ca_commencement_date")
        st.text_input("Hours of Work", key="ca_hours_of_work")
    with c2:
        st.text_input("Contract Rate *", key="ca_contract_rate")
        st.date_input("End Date", key="ca_end_date")
        st.text_input("Notice Period", key="ca_notice_period")
        st.text_input(
            "Other/Travel Expenses",
            value="Upon authorisation by the Nominated Client",
            key="ca_travel_expenses",
        )

    st.divider()
    st.subheader("Output")
    fmt1, fmt2 = st.columns(2)
    with fmt1:
        ca_fmt_docx = st.checkbox(".docx", value=True, key="ca_fmt_docx")
        ca_fmt_pdf = st.checkbox(".pdf", key="ca_fmt_pdf")

    # Auto-save draft
    if ca_client and "ca_generated" not in st.session_state:
        _draft_data = {
            "type": "ltd_company" if is_ltd else "sole_trader",
            "nominated_client": ca_client,
            "role": st.session_state.get("ca_role", ""),
            "commencement_date": str(st.session_state.get("ca_commencement_date", "")),
            "end_date": str(st.session_state.get("ca_end_date", "")),
            "hours_of_work": st.session_state.get("ca_hours_of_work", ""),
            "contract_rate": st.session_state.get("ca_contract_rate", ""),
            "notice_period": st.session_state.get("ca_notice_period", ""),
            "travel_expenses": st.session_state.get("ca_travel_expenses", "Upon authorisation by the Nominated Client"),
        }
        if is_ltd:
            _draft_data.update({
                "provider_company": st.session_state.get("ca_provider_company", ""),
                "trading_as": st.session_state.get("ca_trading_as", ""),
                "registered_address": st.session_state.get("ca_registered_address", ""),
                "company_nzbn": st.session_state.get("ca_company_nzbn", ""),
                "individual_contractor": st.session_state.get("ca_individual_contractor", ""),
                "ird_number": st.session_state.get("ca_ird_number", ""),
                "gst_registered": st.session_state.get("ca_gst_registered", "No"),
                "gst_number": st.session_state.get("ca_gst_number", ""),
                "bank_account": st.session_state.get("ca_bank_account", ""),
            })
        save_draft(user_email, _DOC_TYPE, _draft_data)

    # Generate
    if st.button("Generate Agreement", type="primary", key="ca_generate"):
        missing = []
        if not ca_client:
            missing.append("Nominated Client")
        if not st.session_state.get("ca_role"):
            missing.append("Role")
        if not st.session_state.get("ca_contract_rate"):
            missing.append("Contract Rate")
        if is_ltd and not st.session_state.get("ca_provider_company"):
            missing.append("Provider company name")
        if is_ltd and not st.session_state.get("ca_individual_contractor"):
            missing.append("Individual Contractor name")
        if not ca_fmt_docx and not ca_fmt_pdf:
            missing.append("At least one output format")

        if missing:
            st.error(f"Please fill in: {', '.join(missing)}")
        else:
            _com_date = st.session_state.get("ca_commencement_date", date.today())
            _end_date = st.session_state.get("ca_end_date", date.today())
            _gen_data = {
                "contractor_type": "ltd_company" if is_ltd else "sole_trader",
                "date_of_agreement": f"{date.today().day} {date.today().strftime('%B')} {date.today().year}",
                "nominated_client": ca_client,
                "role": st.session_state.get("ca_role", ""),
                "commencement_date": f"{_com_date.day} {_com_date.strftime('%B')} {_com_date.year}",
                "end_date": f"{_end_date.day} {_end_date.strftime('%B')} {_end_date.year}",
                "hours_of_work": st.session_state.get("ca_hours_of_work", ""),
                "contract_rate": st.session_state.get("ca_contract_rate", ""),
                "notice_period": st.session_state.get("ca_notice_period", ""),
                "travel_expenses": st.session_state.get("ca_travel_expenses", "Upon authorisation by the Nominated Client"),
            }
            if is_ltd:
                _gen_data.update({
                    "provider_company": st.session_state.get("ca_provider_company", ""),
                    "trading_as": st.session_state.get("ca_trading_as", ""),
                    "registered_address": st.session_state.get("ca_registered_address", ""),
                    "company_nzbn": st.session_state.get("ca_company_nzbn", ""),
                    "individual_contractor": st.session_state.get("ca_individual_contractor", ""),
                    "ird_number": st.session_state.get("ca_ird_number", ""),
                    "gst_registered": st.session_state.get("ca_gst_registered", "No") == "Yes",
                    "gst_number": st.session_state.get("ca_gst_number", ""),
                    "bank_account": st.session_state.get("ca_bank_account", ""),
                })
            try:
                docx_bytes = generate_ca(_gen_data)
                st.session_state["ca_generated"] = docx_bytes
                st.session_state["ca_gen_data"] = _gen_data
                delete_draft(user_email, _DOC_TYPE)
                st.rerun()
            except Exception as e:
                st.error(f"Error generating document: {e}")

    # --- Download ---
    if "ca_generated" in st.session_state:
        st.divider()
        st.subheader("Download")
        _gen_data = st.session_state["ca_gen_data"]
        _docx = st.session_state["ca_generated"]
        _ctype_label = "Sole Trader" if _gen_data["contractor_type"] == "sole_trader" else "Ltd Company"
        _fname = f"Contractor Agreement - {_ctype_label} - {_gen_data['nominated_client']}"

        if ca_fmt_docx:
            st.download_button(
                "Download .docx", _docx, f"{_fname}.docx",
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="ca_dl_docx",
            )
        if ca_fmt_pdf:
            _pdf = convert_docx_to_pdf(_docx)
            if _pdf:
                st.download_button(
                    "Download .pdf", _pdf, f"{_fname}.pdf",
                    mime="application/pdf",
                    key="ca_dl_pdf",
                )
            else:
                st.warning("PDF conversion failed.")

else:
    st.info("This document type is coming soon.")
